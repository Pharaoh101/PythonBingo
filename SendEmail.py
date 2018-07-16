import openpyxl
import win32com.client as win32
from NumberGenerator import NumberGenerator
import datetime
import time
import sched
import xlrd
from validate_email import validate_email
import sys

class Email:

    def __init__(self):
        self.book = 0
        self.emails = ""

        # Generate the numbers and insert them into the numbers array in the order we want them to appear in the game
        num = NumberGenerator()
        num.generate_numbers()
        num.fill_lists()
        num.view_numbers(num.numbers)

        self.numbers = [num.first, num.second, num.third, num.fourth, num.fifth, num.sixth]

    # Get the emails from a spreadsheet and into a string, ready to be sent
    def get_emails(self, path_to_excel):
        self.book = openpyxl.load_workbook(path_to_excel)
        sheet = self.book.active
        email_string = ''
        # Loop through the names in the spreadsheet, tidying them up
        for row in sheet.iter_rows('A{}:A{}'.format(sheet.min_row, sheet.max_row)):
            for cell in row:
                email = cell.value.replace('<', '')
                email = email.replace('>', '')
                email = email.strip()
                if email[-1] != ';':
                    email = email + ';'

                email_string += email

        self.emails = email_string


    def send_email(self, subject, round_number):
        # win32 is an old-ish library, allows the program to interact with any Microsoft apps open e.g. Outlook
        # There is no login here, it will simply use the account that's using the current instance of Outlook to send the emails
        #so the emails will look like they came from you
        outlook = win32.Dispatch('outlook.application')
        mail = outlook.CreateItem(0)
        mail.Bcc = self.emails
        mail.Subject = subject
        # Convoluted way of having the signature picture at the top of the bingo emails
        # You'll have to make the pictures yourself and paste the link to it from your directory here 
        attachment = mail.Attachments.Add("PATH-TO-PICTURE")
        attachment.PropertyAccessor.SetProperty("http://schemas.microsoft.com/mapi/proptag/0x3712001F", "MyId1")
        mail.HTMLBody = "<html><body>Test image <img src=""cid:MyId1""></body></html>"
        #mail.HTMLBody = """
        #                <html>
        #                <head></head>
        #                <body>
        #                    <font color="DarkBlue" size=-1 face="Arial">
        #                    <p style="text-align:center;"><img src="cid:MyId1" alt="Logo"></p>
        #                    <p>Hello everyone,</p>
        #                    <p>
        #                        Just a quick reminder that tomorrow Bingo will be on, raising money for CHARITY-NAME, so don't forget your change!<br/>
        #                        Your sellers on each floor will be as follows:<br/>
        #                        <ul style="list-style-type:disc">
        #                            <li>Top Floor: ENTER-NAME</li>
        #                            <li>First Floor: ENTER-NAME</li>
        #                            <li>Bottom Floor: ENTER-NAME</li>
        #                        </ul>
        #                    </p>
        #                    <p>Good luck to everyone!</p>
        #                    <p>Many thanks,<br/>ENTER-NAME</p>
        #                </body>
        #                </html>
        #                """
        mail.HTMLBody = """
                        <html>
                        <head></head>
                        <body>
                            <font color="DarkBlue" size=-1 face="Arial">
                            <p style="text-align:center;"><img src="cid:MyId1" alt="Logo"></p>
                            {0}
                            <h5>RULES</h5>
                            <ul style="list-style-type:disc">
                                <li>£1 per ticket</li>
                                <liFirst round begins @ 1.30pm, Thursday 5th April</li>
                                <li>1 round every 30 mins</li>
                                <li>13 numbers per round</li>
                                <li>There will be no more than 6 rounds</li>
                                <li>If there are multiple winners in the same round, one winner will be selected on the order of the numbers in the round</li>
                                <li>If no winner is found after the 6 rounds, there will be a rollover to the next week</li>
                                <li>You have half an hour between rounds to check your ticket and call BINGO</li>
                                <li>If you have BINGO and fail to call it before the next round starts, you may lose your chance to claim your winning ticket</li>
                                <li>You need a full house to win. (EVERY number on your ticket must have been called)</li>
                                <li>Reply back to this email address with “BINGO” as soon as you get a full house</li>
                            </ul>
                        </font>
                        </body>
                        </html>
                        """.format(self.format_numbers(round_number))
        mail.Send()

    # Format the numbers to fit in with the email
    def format_numbers(self, round):
        email_numbers = ""
        for i in range(0, round):
            round_numbers = str(self.numbers[i])
            round_numbers = round_numbers.replace(", ", " - ")
            round_numbers = round_numbers.replace("[", "")
            round_numbers = round_numbers.replace("]", "")
            email_numbers = """<p style="text-align:center;color:black;font-size:30px">{0}</p>""".format(round_numbers) + email_numbers
            email_numbers = """<h4 style="text-align:center;text-decoration: underline;">Round {0}</h4>""".format(i + 1) + email_numbers
        return email_numbers

# Initialise the email
email = Email()

# Make sure you have an up-to-date mailing list
email.get_emails("BingoMailingList.xlsx")

# Scheduler class allows you to keep the program running, and every x seconds it will perform an operation
# In this example, every 1800 seconds (30 minutes) an email is sent out to everyone
scheduler = sched.scheduler(time.time, time.sleep)
scheduler.enter(0, 1, email.send_email, ("Bingo Round 1", 1,))
scheduler.enter(1800, 1, email.send_email, ("Bingo Round 2", 2,))
scheduler.enter(3600, 1, email.send_email, ("Bingo Round 3", 3,))
scheduler.enter(5400, 1, email.send_email, ("Bingo Round 4", 4,))
scheduler.enter(7200, 1, email.send_email, ("Bingo Round 5", 5,))
scheduler.enter(9000, 1, email.send_email, ("Bingo Round 6", 6,))
scheduler.run()
#print(email.emails)